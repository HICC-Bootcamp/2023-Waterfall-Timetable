from openxl import Workbook
wb = Workbook()
ws = wb.active
ws_new = wb.create_sheet()
ws.title = "class1"
ws_new.title = "class2"

wb.save("timetable.xlsx")

from openpyxl import load_workbook

wb = load_workbook("timetable.xlsx")

ws = wb["class1"]